import io
from datetime import datetime, UTC
from uuid import UUID
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.collaboration import Approval, AuditLog
from app.models.decision import (
    Alternative,
    AlternativeEvaluation,
    Decision,
    EvaluationCriterion,
    Risk,
    Stakeholder,
)
from app.models.identity import Role, Team, User, UserProfile
from app.models.taxonomy import DecisionCategory
from app.services.decision_service import calculate_alternative_scores


def generate_decision_pdf(db: Session, decision_id: UUID) -> io.BytesIO:
    """Generate a formal enterprise Decision Record PDF."""
    decision = db.scalar(select(Decision).where(Decision.id == decision_id))
    if not decision:
        raise ValueError("Decision not found")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "DocTitle",
        parent=styles["Heading1"],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=12,
    )
    h2_style = ParagraphStyle(
        "DocH2",
        parent=styles["Heading2"],
        fontSize=13,
        leading=16,
        textColor=colors.HexColor("#1e3a8a"),
        spaceBefore=12,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "DocBody",
        parent=styles["Normal"],
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#334155"),
    )
    bold_body = ParagraphStyle(
        "DocBold",
        parent=body_style,
        fontName="Helvetica-Bold",
    )

    story = []

    # Title & Header
    story.append(Paragraph("EXPERT DECISION REPLAY PLATFORM", ParagraphStyle("Header", fontSize=9, textColor=colors.HexColor("#64748b"), spaceAfter=4)))
    story.append(Paragraph(f"Decision Record: {decision.title}", title_style))
    story.append(Spacer(1, 8))

    # Meta Table
    owner = db.scalar(select(User).where(User.id == decision.owner_id))
    owner_profile = db.scalar(select(UserProfile).where(UserProfile.user_id == decision.owner_id)) if owner else None
    owner_name = owner_profile.full_name if owner_profile else (owner.email if owner else "N/A")

    cat_name = "None"
    if decision.category_id:
        c = db.scalar(select(DecisionCategory).where(DecisionCategory.id == decision.category_id))
        if c:
            cat_name = c.name

    meta_data = [
        [Paragraph("<b>Status:</b>", body_style), Paragraph(decision.status.upper(), bold_body), Paragraph("<b>Version:</b>", body_style), Paragraph(f"v{decision.current_version_no}", body_style)],
        [Paragraph("<b>Owner:</b>", body_style), Paragraph(owner_name, body_style), Paragraph("<b>Category:</b>", body_style), Paragraph(cat_name, body_style)],
        [Paragraph("<b>Implementation:</b>", body_style), Paragraph(decision.implementation_status.upper(), body_style), Paragraph("<b>Created:</b>", body_style), Paragraph(decision.created_at.strftime("%Y-%m-%d %H:%M"), body_style)],
    ]
    t_meta = Table(meta_data, colWidths=[90, 180, 80, 180])
    t_meta.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f8fafc')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#cbd5e1')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t_meta)
    story.append(Spacer(1, 14))

    # Problem Statement
    story.append(Paragraph("1. Problem Statement & Strategic Context", h2_style))
    story.append(Paragraph(decision.problem_statement.replace("\n", "<br/>"), body_style))
    story.append(Spacer(1, 12))

    # Alternatives & Scores
    story.append(Paragraph("2. Evaluated Alternatives & Comparative Analysis", h2_style))
    scores = calculate_alternative_scores(db, decision.id)
    alts = db.scalars(
        select(Alternative)
        .where(Alternative.decision_id == decision.id, Alternative.deleted_at.is_(None))
        .order_by(Alternative.sort_order)
    ).all()

    if alts:
        alt_table_data = [[
            Paragraph("<b>Alternative</b>", bold_body),
            Paragraph("<b>Description</b>", bold_body),
            Paragraph("<b>Score</b>", bold_body),
            Paragraph("<b>Selected</b>", bold_body),
        ]]
        for a in alts:
            sc = str(scores.get(a.id, "N/A"))
            is_sel = "YES (SELECTED)" if (a.is_selected or a.id == decision.selected_alternative_id) else "No"
            alt_table_data.append([
                Paragraph(a.title, bold_body if is_sel.startswith("YES") else body_style),
                Paragraph(a.description or "?", body_style),
                Paragraph(sc, bold_body),
                Paragraph(is_sel, bold_body if is_sel.startswith("YES") else body_style),
            ])
        t_alts = Table(alt_table_data, colWidths=[130, 260, 60, 80])
        t_alts.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(t_alts)
    else:
        story.append(Paragraph("No alternatives recorded.", body_style))
    story.append(Spacer(1, 12))

    # Risks & Mitigations
    story.append(Paragraph("3. Risk Assessment & Mitigations", h2_style))
    risks = db.scalars(
        select(Risk).where(Risk.decision_id == decision.id, Risk.deleted_at.is_(None))
    ).all()
    if risks:
        risk_data = [[
            Paragraph("<b>Risk</b>", bold_body),
            Paragraph("<b>Severity</b>", bold_body),
            Paragraph("<b>Likelihood</b>", bold_body),
            Paragraph("<b>Mitigation Strategy</b>", bold_body),
        ]]
        for r in risks:
            risk_data.append([
                Paragraph(r.title, bold_body),
                Paragraph(r.severity.upper(), body_style),
                Paragraph(r.likelihood.upper(), body_style),
                Paragraph(r.mitigation or "?", body_style),
            ])
        t_risks = Table(risk_data, colWidths=[150, 65, 65, 250])
        t_risks.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(t_risks)
    else:
        story.append(Paragraph("No risks recorded.", body_style))
    story.append(Spacer(1, 12))

    # Approvals Timeline
    story.append(Paragraph("4. Approval & Audit Sign-offs", h2_style))
    approvals = db.scalars(
        select(Approval).where(Approval.decision_id == decision.id).order_by(Approval.step_order)
    ).all()
    if approvals:
        appr_data = [[
            Paragraph("<b>Step</b>", bold_body),
            Paragraph("<b>Sign-off Name</b>", bold_body),
            Paragraph("<b>Status</b>", bold_body),
            Paragraph("<b>Actor / Approver</b>", bold_body),
            Paragraph("<b>Date / Comments</b>", bold_body),
        ]]
        for ap in approvals:
            actor = db.scalar(select(User).where(User.id == ap.actor_id)) if ap.actor_id else None
            actor_name = actor.email if actor else "?"
            acted_date = ap.acted_at.strftime("%Y-%m-%d %H:%M") if ap.acted_at else "Pending"
            appr_data.append([
                Paragraph(f"Step {ap.step_order}", body_style),
                Paragraph(ap.step_name, bold_body),
                Paragraph(ap.status.upper(), bold_body),
                Paragraph(actor_name, body_style),
                Paragraph(f"{acted_date}<br/><i>{ap.comment or ''}</i>", body_style),
            ])
        t_appr = Table(appr_data, colWidths=[45, 145, 75, 115, 150])
        t_appr.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(t_appr)
    else:
        story.append(Paragraph("No approval records available.", body_style))
    story.append(Spacer(1, 12))

    # Outcome
    if decision.outcome_summary:
        story.append(Paragraph("5. Decision Outcome & Retrospective Rationale", h2_style))
        story.append(Paragraph(decision.outcome_summary.replace("\n", "<br/>"), body_style))

    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_decision_excel(db: Session, decision_id: UUID) -> io.BytesIO:
    """Generate a multi-tab enterprise decision workbook in Excel."""
    decision = db.scalar(select(Decision).where(Decision.id == decision_id))
    if not decision:
        raise ValueError("Decision not found")

    wb = openpyxl.Workbook()
    
    # Sheet 1: Overview
    ws_meta = wb.active
    ws_meta.title = "Decision Overview"
    ws_meta.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14, color="0F172A")
    bold_font = Font(bold=True, size=10)

    ws_meta.cell(row=1, column=1, value="EXPERT DECISION REPLAY PLATFORM ? CASE FILE").font = title_font
    ws_meta.cell(row=3, column=1, value="Title").font = bold_font
    ws_meta.cell(row=3, column=2, value=decision.title)
    ws_meta.cell(row=4, column=1, value="Status").font = bold_font
    ws_meta.cell(row=4, column=2, value=decision.status.upper())
    ws_meta.cell(row=5, column=1, value="Implementation Status").font = bold_font
    ws_meta.cell(row=5, column=2, value=decision.implementation_status.upper())
    ws_meta.cell(row=6, column=1, value="Version").font = bold_font
    ws_meta.cell(row=6, column=2, value=f"v{decision.current_version_no}")
    ws_meta.cell(row=7, column=1, value="Created At").font = bold_font
    ws_meta.cell(row=7, column=2, value=decision.created_at.strftime("%Y-%m-%d %H:%M"))
    ws_meta.cell(row=9, column=1, value="Problem Statement").font = bold_font
    ws_meta.cell(row=9, column=2, value=decision.problem_statement)
    if decision.outcome_summary:
        ws_meta.cell(row=11, column=1, value="Outcome Rationale").font = bold_font
        ws_meta.cell(row=11, column=2, value=decision.outcome_summary)

    # Sheet 2: Alternatives & Scoring
    ws_alts = wb.create_sheet(title="Alternatives & Scoring")
    ws_alts.views.sheetView[0].showGridLines = True
    alt_headers = ["Title", "Description", "Score", "Selected"]
    for col_idx, h in enumerate(alt_headers, start=1):
        c = ws_alts.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font

    scores = calculate_alternative_scores(db, decision.id)
    alts = db.scalars(
        select(Alternative).where(Alternative.decision_id == decision.id, Alternative.deleted_at.is_(None)).order_by(Alternative.sort_order)
    ).all()
    for row_idx, a in enumerate(alts, start=2):
        ws_alts.cell(row=row_idx, column=1, value=a.title)
        ws_alts.cell(row=row_idx, column=2, value=a.description or "")
        ws_alts.cell(row=row_idx, column=3, value=scores.get(a.id, "N/A"))
        ws_alts.cell(row=row_idx, column=4, value="YES" if (a.is_selected or a.id == decision.selected_alternative_id) else "NO")

    # Sheet 3: Risks
    ws_risks = wb.create_sheet(title="Risks & Mitigations")
    ws_risks.views.sheetView[0].showGridLines = True
    risk_headers = ["Risk Title", "Severity", "Likelihood", "Mitigation Strategy"]
    for col_idx, h in enumerate(risk_headers, start=1):
        c = ws_risks.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font

    risks = db.scalars(select(Risk).where(Risk.decision_id == decision.id, Risk.deleted_at.is_(None))).all()
    for row_idx, r in enumerate(risks, start=2):
        ws_risks.cell(row=row_idx, column=1, value=r.title)
        ws_risks.cell(row=row_idx, column=2, value=r.severity.upper())
        ws_risks.cell(row=row_idx, column=3, value=r.likelihood.upper())
        ws_risks.cell(row=row_idx, column=4, value=r.mitigation or "")

    # Sheet 4: Approval Trail
    ws_appr = wb.create_sheet(title="Approval History")
    ws_appr.views.sheetView[0].showGridLines = True
    appr_headers = ["Step", "Step Name", "Status", "Acted By", "Acted At", "Comments"]
    for col_idx, h in enumerate(appr_headers, start=1):
        c = ws_appr.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font

    approvals = db.scalars(select(Approval).where(Approval.decision_id == decision.id).order_by(Approval.step_order)).all()
    for row_idx, ap in enumerate(approvals, start=2):
        actor = db.scalar(select(User).where(User.id == ap.actor_id)) if ap.actor_id else None
        ws_appr.cell(row=row_idx, column=1, value=f"Step {ap.step_order}")
        ws_appr.cell(row=row_idx, column=2, value=ap.step_name)
        ws_appr.cell(row=row_idx, column=3, value=ap.status.upper())
        ws_appr.cell(row=row_idx, column=4, value=actor.email if actor else "?")
        ws_appr.cell(row=row_idx, column=5, value=ap.acted_at.strftime("%Y-%m-%d %H:%M") if ap.acted_at else "Pending")
        ws_appr.cell(row=row_idx, column=6, value=ap.comment or "")

    # Adjust widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_summary_excel(db: Session) -> io.BytesIO:
    """Generate an organization-wide summary report of all decisions in Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Decisions Summary"
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    headers = ["Title", "Category", "Status", "Implementation", "Owner", "Version", "Created At"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font

    decisions = db.scalars(select(Decision).where(Decision.deleted_at.is_(None)).order_by(Decision.created_at.desc())).all()
    for row_idx, d in enumerate(decisions, start=2):
        owner = db.scalar(select(User).where(User.id == d.owner_id))
        cat = db.scalar(select(DecisionCategory).where(DecisionCategory.id == d.category_id)) if d.category_id else None
        ws.cell(row=row_idx, column=1, value=d.title)
        ws.cell(row=row_idx, column=2, value=cat.name if cat else "Uncategorized")
        ws.cell(row=row_idx, column=3, value=d.status.upper())
        ws.cell(row=row_idx, column=4, value=d.implementation_status.upper())
        ws.cell(row=row_idx, column=5, value=owner.email if owner else "?")
        ws.cell(row=row_idx, column=6, value=f"v{d.current_version_no}")
        ws.cell(row=row_idx, column=7, value=d.created_at.strftime("%Y-%m-%d"))

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
